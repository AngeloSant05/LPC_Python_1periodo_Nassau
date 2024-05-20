import openpyxl

# Utilizar uma planilha já existente
book = openpyxl.load_workbook('Planilho de entrada e saida de caixa.xlsx')

# Selecionando uma página
fluxo = book['Fluxo']

# Imprimindo os dados de cada linha
# for rows in fluxo.iter_rows(min_row=2):
    # print(rows[0].value, rows[1].value, rows[2].value)
    # Ou
    # print(f'{rows[0].value}, {rows[1].value}, {rows[2].value}')
    # Ou
    # for cell in rows:
        # print(cell.value)

# Alterando valores
    # for cell in rows:
        # if cell.value == 'mes':
            # cell.value == 'jan'
             
# Renomeando uma página
Sheet = book['Sheet']
Sheet.title = 'Pera'
print(book.sheetnames)

# Deletando uma página
del book['Pera']
print(book.sheetnames)

# Salvando as alterações
book.save('Planilha entrada e saida de caixa1.xlsx')

