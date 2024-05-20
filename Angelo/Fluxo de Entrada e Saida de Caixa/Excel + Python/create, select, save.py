import openpyxl

# Criando uma planilha (livro/book)
book = openpyxl.Workbook()
# Como visualizar páginas existentes
print(book.sheetnames)
# Como criar uma página
book.create_sheet('Fluxo')
# Como selecionar uma página
fluxo = book['Fluxo']
# Como adicionar coisas na página
fluxo.append(['mes', 'ano', 'valor'])
fluxo.append(['mes', 'ano', 'valor'])
fluxo.append(['mes', 'ano', 'valor'])
fluxo.append(['mes', 'ano', 'valor'])
fluxo.append(['mes', 'ano', 'valor'])
# Salvar a planilha
book.save('Planilho de entrada e saida de caixa.xlsx')

